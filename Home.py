import streamlit as st
import random
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import tempfile

# -------------------------
# CONFIG
# -------------------------
OTP_EXPIRY_SECONDS = 300  # 5 minutes

# -------------------------
# SECRETS
# -------------------------
GMAIL_USER = st.secrets["GMAIL_USER"]
GMAIL_APP_PASSWORD = st.secrets["GMAIL_APP_PASSWORD"]
ALLOWED_EMAIL = st.secrets["ALLOWED_EMAIL"]

# -------------------------
# SEND OTP FUNCTION
# -------------------------
def send_otp_gmail(receiver_email, otp):
    msg = MIMEMultipart()
    msg["From"] = GMAIL_USER
    msg["To"] = receiver_email
    msg["Subject"] = "🔐 Your Streamlit Login OTP"

    body = f"""
    <h2>Your OTP is: {otp}</h2>
    <p>This OTP is valid for <b>5 minutes</b>.</p>
    <p>Do not share it with anyone.</p>
    """
    msg.attach(MIMEText(body, "html"))

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_USER, receiver_email, msg.as_string())
        server.quit()
        return True
    except Exception as e:
        st.error(f"Email send failed: {e}")
        return False

# -------------------------
# OTP AUTH
# -------------------------
if "auth_ok" not in st.session_state:
    st.session_state.auth_ok = False
if "otp_sent" not in st.session_state:
    st.session_state.otp_sent = False
if "otp_time" not in st.session_state:
    st.session_state.otp_time = None
if "otp" not in st.session_state:
    st.session_state.otp = ""

st.title("📊 Secure Streamlit App")
st.subheader("🔐 Gmail OTP Login")

email = st.text_input("Enter your Gmail address", key="email_input")

if st.button("📨 Send OTP"):
    if email.strip().lower() != ALLOWED_EMAIL.lower():
        st.error("❌ Unauthorized email")
    else:
        otp = str(random.randint(100000, 999999))
        st.session_state.otp = otp
        st.session_state.otp_time = time.time()
        st.session_state.otp_sent = send_otp_gmail(email, otp)
        if st.session_state.otp_sent:
            st.success("✅ OTP sent to your Gmail")

entered_otp = st.text_input("Enter OTP", max_chars=6, key="otp_input")
if st.button("✅ Verify OTP"):
    if not st.session_state.otp_sent:
        st.warning("⚠️ Please request OTP first")
    elif time.time() - st.session_state.otp_time > OTP_EXPIRY_SECONDS:
        st.error("⏰ OTP expired. Send a new one.")
        st.session_state.otp_sent = False
    elif entered_otp == st.session_state.otp:
        st.session_state.auth_ok = True
        st.success("🎉 Login successful")
    else:
        st.error("❌ Invalid OTP")

# -------------------------
# Excel Comparison Section
# -------------------------
if st.session_state.auth_ok:
    st.subheader("📂 Excel Comparison (R0 vs R1)")

    if "r0_file" not in st.session_state:
        st.session_state.r0_file = None
    if "r1_file" not in st.session_state:
        st.session_state.r1_file = None

    col1, col2 = st.columns(2)
    with col1:
        uploaded_r0 = st.file_uploader("Upload R0.xlsx", type=["xlsx"], key="r0_uploader")
        if uploaded_r0:
            st.session_state.r0_file = uploaded_r0
    with col2:
        uploaded_r1 = st.file_uploader("Upload R1.xlsx", type=["xlsx"], key="r1_uploader")
        if uploaded_r1:
            st.session_state.r1_file = uploaded_r1

    r0_file = st.session_state.r0_file
    r1_file = st.session_state.r1_file

    run_btn = st.button("▶️ Run Comparison")

    if run_btn:
        if not r0_file or not r1_file:
            st.warning("⚠️ Please upload both R0.xlsx and R1.xlsx")
        elif r0_file.name != "R0.xlsx" or r1_file.name != "R1.xlsx":
            st.error("❌ Uploaded files must be named exactly R0.xlsx and R1.xlsx")
        else:
            try:
                r0_df = pd.read_excel(r0_file, dtype=str).fillna("")
                r1_df = pd.read_excel(r1_file, dtype=str).fillna("")

                r0_df = r0_df.drop_duplicates(subset="Tag").set_index("Tag")
                r1_df = r1_df.drop_duplicates(subset="Tag").set_index("Tag")

                r0_columns = list(r0_df.columns)
                all_columns = sorted(
                    set(r0_df.columns).union(set(r1_df.columns)),
                    key=lambda x: (r0_columns.index(x) if x in r0_columns else float("inf"))
                )

                all_tags = sorted(set(r0_df.index).union(set(r1_df.index)))
                comparison_rows = []

                for tag in all_tags:
                    if tag not in r0_df.index:
                        row = {"Tag": tag, "Change_Type": "✅ Added in R1"}
                        row.update({col: r1_df.loc[tag].get(col, "") for col in all_columns})
                        row["Change_Summary"] = ""
                        comparison_rows.append(row)
                    elif tag not in r1_df.index:
                        row = {"Tag": tag, "Change_Type": "❌ Removed in R1"}
                        row.update({col: r0_df.loc[tag].get(col, "") for col in all_columns})
                        row["Change_Summary"] = ""
                        comparison_rows.append(row)
                    else:
                        row_r0 = r0_df.loc[tag]
                        row_r1 = r1_df.loc[tag]
                        row_data = {"Tag": tag}
                        summary = []
                        changes_exist = False
                        for col in all_columns:
                            val_r0 = row_r0.get(col, "")
                            val_r1 = row_r1.get(col, "")
                            if str(val_r0).strip() != str(val_r1).strip():
                                row_data[col] = f"{val_r0} → {val_r1}"
                                summary.append(f"{col}: {val_r0} → {val_r1}")
                                changes_exist = True
                            else:
                                row_data[col] = val_r1
                        row_data["Change_Type"] = "✏️ Modified" if changes_exist else "No Change"
                        row_data["Change_Summary"] = " | ".join(summary) if summary else ""
                        comparison_rows.append(row_data)

                comparison_df = pd.DataFrame(comparison_rows)
                final_columns = ["Tag", "Change_Type"] + all_columns + ["Change_Summary"]
                comparison_df = comparison_df[final_columns]

                # Save Excel with highlights
                wb = Workbook()
                ws = wb.active
                ws.title = "Comparison Summary"
                highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                for r_idx, row in enumerate(dataframe_to_rows(comparison_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if isinstance(value, str) and "→" in value:
                            cell.fill = highlight

                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                wb.save(temp_file.name)

                with open(temp_file.name, "rb") as f:
                    st.download_button(
                        "⬇️ Download Excel Comparison",
                        f,
                        file_name=f"Vimal_Comparison_R0_vs_R1_{datetime.now().strftime('%d_%m_%Y_%H_%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                st.success("✅ Comparison completed successfully")
                st.info("🙏 Thank You Vimal")

            except Exception as e:
                st.error(f"⚠️ Error: {e}")
